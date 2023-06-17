import re
import openpyxl
import pandas as pd
import datetime
import random
import string
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from tkinter import ttk
from Clases.Tarjeta import Tarjeta



def generar_archivo():
    try:
        # Obtener la ubicacion del archivo de entrada
        ruta_archivo = entrada_archivo.get()

        if not ruta_archivo:
            raise ValueError("No se selecciono un archivo de entrada.")

        # Busqueda de la info en el Excel
        datos = pd.read_excel(ruta_archivo, usecols=[0], header=None, engine="openpyxl")
        columna = datos[0]
        listaExcel = columna.dropna().astype(str).tolist()

        listaDeTarjetas = []
        for numero in listaExcel:
            numeros_sin_espacios_ni_cero = numero.replace(' ', '').replace('.0', '')
            tarjeta = Tarjeta(numeros_sin_espacios_ni_cero)
            listaDeTarjetas.append(tarjeta)


        # Crear un libro y una hoja de calculo con openpyxl
        workbook = Workbook()
        worksheet = workbook.active

        # Crear el DataFrame
        data = []
        for tarjeta in listaDeTarjetas:
            data.append([
                "'"+tarjeta.numero,
                tarjeta.ValidarTarjeta()
            ])

        df = pd.DataFrame(data, columns=['Tarjeta', 'Resultado'])

        # Escribir los datos en la hoja de calculo
        for row_num, row_data in enumerate(df.values, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num, value=value)

        # Configurar el formato y estilo
        header_fill = PatternFill(fill_type='solid', fgColor='000000')
        header_font = Font(color='FFFFFF', bold=True)
        data_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))
        gray_fill = PatternFill(fill_type='solid', fgColor='E9E9E9')

        # Aplicar formato y estilo a los encabezados
        for col_num, header in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Aplicar formato y estilo a las celdas de datos
        for row in worksheet.iter_rows(min_row=2, min_col=1, max_col=df.shape[1]):
            for cell in row:
                cell.border = data_border

        # Aplicar formato y estilo a las filas alternas
        for row_num in range(2, worksheet.max_row + 1, 2):
            for col_num in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.fill = gray_fill

        # Ajustar el ancho de las columnas
        for column in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

        # Generar un nombre de archivo aleatorio basado en la fecha y hora
        timestamp = datetime.datetime.now().strftime("%Y%m%d")
        random_word = ''.join(random.choices(string.ascii_lowercase, k=5))
        nombre_archivo = f"{timestamp}_{random_word}.xlsx"

        # Ruta de guardado con el nombre de archivo generado
        ruta_guardado = filedialog.asksaveasfilename(defaultextension=".xlsx")

        if not ruta_guardado:
            raise ValueError("No se selecciono una ubicacion de guardado.")

        # Guardar el archivo Excel
        workbook.save(ruta_guardado)
        workbook.close()

        # Mostrar la ubicacion del archivo generado
        resultado_archivo.config(text="Archivo generado:\n" + ruta_guardado)
        messagebox.showinfo("Proceso completado", "El archivo se ha generado correctamente.")

    except Exception as e:
        messagebox.showerror("Error", str(e))


# Funcion para seleccionar un archivo
def seleccionar_archivo():
    archivo = filedialog.askopenfilename(filetypes=[("Archivos Excel", "*.xlsx")])
    entrada_archivo.delete(0, tk.END)
    entrada_archivo.insert(0, archivo)


# Crear la interfaz de usuario

window = tk.Tk()
window.title("Generador de Archivo Excel")

style = ttk.Style()
style.configure("TButton", font=("Arial", 12), padding=6)

# Etiqueta para mostrar el texto "Seleccione el Excel"
lbl_archivo = ttk.Label(window, text="Seleccione el Excel:", font=("Arial", 12, "bold"))
lbl_archivo.pack(pady=10)

# Marco que contiene el campo de entrada y el boton "Seleccionar"
frame_seleccionar = ttk.Frame(window)
frame_seleccionar.pack(pady=5)

# Campo de entrada deshabilitado para mostrar la ubicacion del archivo seleccionado
entrada_archivo = ttk.Entry(frame_seleccionar, width=50)
entrada_archivo.pack(side=tk.LEFT, padx=5)

# Boton "Seleccionar" para elegir un archivo
btn_seleccionar = ttk.Button(frame_seleccionar, text="Seleccionar", command=seleccionar_archivo)
btn_seleccionar.pack(side=tk.LEFT, padx=5)

# Boton "Generar" para generar el archivo Excel
btn_generar = ttk.Button(window, text="Generar", style="TButton", command=generar_archivo)
btn_generar.pack(pady=10)

# Etiqueta para mostrar la ubicacion del archivo generado
resultado_archivo = ttk.Label(window, text="", font=("Arial", 12))
resultado_archivo.pack()

window.mainloop()






   



